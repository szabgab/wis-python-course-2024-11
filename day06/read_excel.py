def read_excelwith_openpyxl(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end=' ')
        print()
    wb.close()

def read_excel(file_path):
    import pandas as pd
    df = pd.read_excel(file_path)
    print(df)
    df.to_excel('output.xlsx', index=False)
